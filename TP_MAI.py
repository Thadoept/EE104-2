# -*- coding: utf-8 -*-
"""
Created on Sat Sep 12 15:32:45 2020

@author: Thadoe
"""

### Create Cashier ###

#Read an excel file using python

import pandas
import numpy 
import json
import re
from openpyxl import load_workbook
import smtplib


print ("Welcome to my convenience store!")
cont = True

### Cash Inventory ###
def CashInventory(total_bill):
    print("\n Your total today is {0}".format(total_bill))
    paid = float(input()) 
    bank = {100.0:0, 50.0:0, 20.0:0, 10.0:0, 5.0:0, 1.0:0, 0.25:0, 0.1:0, 0.01:0}
    due = paid - total_bill
    if paid<total_bill:
        print("Sorry! It's not enough money")
    else:
        for change in sorted(bank, reverse=True):
            amt = max(0, due//change)
            due = due-(amt*change)
            bank[change] = int(amt)
        print(bank)
    
### Calculator ###
def Calculator(chosen_item_list, qty_list, price_list):
    total_bill = 0
    for n in range(0, len(qty_list)):
        bill = price_list[n]*qty_list[n]
        total_bill = total_bill + bill
    CashInventory(total_bill)
       
    
def GoBack():
    anything_else = str(input("Cashier: Do you want anything else?: Yes or No? \nCustomer: "))  
    if anything_else == 'No':
       global cont 
       cont = False
       Calculator(chosen_item_list, qty_list, price_list)
       return cont
   
def getPrice(chosen_item): #Get the price of the product from excel
    df = pandas.read_excel (r'C:\Users\Thadoe\Desktop\EE104 (TPT)\Project 1\store.xlsx',index_col ="Product")
    price=df.loc[chosen_item,"Price"] 
    return(price) 

### Cart ### 
chosen_item_list = []
qty_list = [] 
price_list = []

def Cart(chosen_item, qty, c):
    chosen_item_list.append(chosen_item)
    qty_list.append(qty)
    price_list.append(c)
    #print (chosen_item_list, qty_list, price_list)
    GoBack() 

    
while cont == True: 
    menu = pandas.read_excel(r'C:\Users\Thadoe\Desktop\EE104 (TPT)\Project 1\store.xlsx', sheet_name='menu')
    print(menu)
    ask_item = input("Cashier: What would you like to purchase today? \n" + "Customer: ")
    ask_qty = input("Cashier: How many do you want? \n" + "Customer:")
   
    
    chosen_item = str(ask_item)
    qty = int(ask_qty)   
    c = float(getPrice(chosen_item))
    
    ### Store Management / stock ###
    items_instock = menu['Product'].tolist()
    qty_instock = menu['Qty'].tolist()
    
    
    if chosen_item in items_instock:
        item_index = items_instock.index(chosen_item)
        qty_index = int(qty_instock[item_index])
        
        qty_left = qty_index - qty
        print(qty_left)
        
        if qty_left<5:
            sender_address = "ee140project1sjsu@gmail.com" # Replace this with your Gmail address

            receiver_address = "thadoepyaethu@gmail.com" # Replace this with any valid email address

            account_password = "sjsusjsu" # Replace this with your Gmail account password

            subject = "Out of stock alert"

            body = ("The stock of {0} is getting low!!!".format (chosen_item))

            # Endpoint for the SMTP Gmail server (Don't change this!)
            smtp_server = smtplib.SMTP_SSL("smtp.gmail.com", 465)

            # Login with your Gmail account using SMTP
            smtp_server.login(sender_address, account_password)
            
            # Let's combine the subject and the body onto a single message
            message = f"Subject: {subject}\n\n{body}"
            
            # We'll be sending this message in the above format (Subject:...\n\nBody)
            smtp_server.sendmail(sender_address, receiver_address, message)
            
            # Close our endpoint
            smtp_server.close()

        
        # wb = load_workbook("C:/Users/Thadoe/Desktop/EE104 (TPT)/Project 1/store.xlsx")
        # ws = wb["menu"]
        
        # wcell1 = ws.cell(row=int(item_index),column=3)
        # wcell1.value = int(qty_left)
        
        # wb.save("C:/Users/Thadoe/Desktop/EE104 (TPT)/Project 1/store.xlsx")
        
        if qty < qty_index:
            
            Cart(chosen_item, qty, c)
            
        else:
            print("Cashier: Out of stock")
            GoBack()
    else:
        print("Cashier: Sorry! it is not available.\n")
        GoBack()

    

     
    

